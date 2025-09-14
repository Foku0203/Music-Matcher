# matcher/management/commands/import_workbook.py
from django.core.management.base import BaseCommand, CommandError
from django.db import transaction
from django.utils import timezone
from django.utils.dateparse import parse_datetime, parse_date

from openpyxl import load_workbook

from matcher.models import AppUser, Artist, Album, Genre, Song, SongGenre

# Optional (มีค่อย import)
try:
    from matcher.models import Emotion, SongEmotion
    HAS_EMOTION = True
except Exception:
    HAS_EMOTION = False


# -------------------- utils --------------------
def norm_key(s: str) -> str:
    return "".join(c for c in (str(s or "")).lower() if c.isalnum())

def pick(row: dict, candidates, default=""):
    """หยิบค่าแรกที่พบตามชื่อคอลัมน์ที่เป็นไปได้ (ยืดหยุ่น)"""
    keys = {norm_key(k): k for k in row.keys()}
    for cand in candidates:
        k = keys.get(norm_key(cand))
        if k in row:
            v = row.get(k)
            if v is None:
                return default
            return str(v).strip() if not isinstance(v, (int, float)) else str(v)
    return default

def as_int(val, default=None):
    try:
        if val is None or val == "":
            return default
        return int(float(val))
    except Exception:
        return default

def as_bool(val, default=None):
    if val is None or val == "":
        return default
    s = str(val).strip().lower()
    if s in {"1","true","t","yes","y"}: return True
    if s in {"0","false","f","no","n"}: return False
    return default

def parse_dt(val):
    """รองรับค่าจาก Excel (datetime) หรือสตริง"""
    if not val:
        return timezone.now()
    if hasattr(val, "year") and hasattr(val, "month") and hasattr(val, "day"):
        dt = val
    else:
        dt = parse_datetime(str(val))
        if not dt:
            d = parse_date(str(val))
            if d:
                from datetime import datetime
                dt = datetime(d.year, d.month, d.day)
    if not dt:
        return timezone.now()
    return dt if timezone.is_aware(dt) else timezone.make_aware(dt)

def sheet_dicts(ws):
    rows = ws.iter_rows(values_only=True)
    try:
        headers = [str(h).strip() if h is not None else "" for h in next(rows)]
    except StopIteration:
        return []
    for r in rows:
        yield {headers[i]: (r[i] if i < len(r) else None) for i in range(len(headers))}


# -------------------- helpers สำหรับ users --------------------
from uuid import UUID

def find_existing_user(uid_str, email, username):
    """พยายามหา user เดิม: user_id -> email -> username"""
    if uid_str:
        try:
            return AppUser.objects.get(user_id=UUID(str(uid_str)))
        except Exception:
            pass
    if email:
        try:
            return AppUser.objects.get(email=email)
        except AppUser.DoesNotExist:
            pass
    if username:
        try:
            return AppUser.objects.get(username=username)
        except AppUser.DoesNotExist:
            pass
    return None

def unique_username(desired, exclude_id=None):
    """คืนค่า username ที่ไม่ซ้ำ: alice -> alice_2 -> alice_3 ..."""
    base = (desired or "user").strip() or "user"
    uname = base
    i = 2
    qs = AppUser.objects.all()
    if exclude_id:
        qs = qs.exclude(user_id=exclude_id)
    while qs.filter(username=uname).exists():
        uname = f"{base}_{i}"
        i += 1
    return uname


# -------------------- importers --------------------
def import_users(ws):
    seen = created = updated = renamed = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        seen += 1
        uid_str    = pick(row, ["user_id","id","uuid"])
        email      = pick(row, ["email"])
        username   = pick(row, ["username","user","login","name"])
        pwd_hash   = pick(row, ["password_hash","password","passhash"])
        status     = pick(row, ["status","state"], "active").lower()
        created_at = parse_dt(pick(row, ["created_at","created","joined"]))

        if not email or not username:
            raise CommandError(f"[users] row {i}: ต้องมี email และ username")

        status = status if status in {"active","suspended"} else "active"
        existing = find_existing_user(uid_str, email, username)

        if existing:
            old_username = existing.username
            new_username = username if username == old_username else unique_username(username, exclude_id=existing.user_id)
            if new_username != old_username:
                renamed += 1
            existing.email = email or existing.email
            existing.username = new_username
            if pwd_hash:
                existing.password_hash = pwd_hash
            existing.status = status
            if created_at:
                existing.created_at = created_at
            existing.save()
            updated += 1
        else:
            final_username = unique_username(username)
            defaults = {
                "email": email,
                "username": final_username,
                "password_hash": pwd_hash,
                "status": status,
                "created_at": created_at,
            }
            if uid_str:
                try:
                    uid = UUID(uid_str)
                except Exception:
                    raise CommandError(f"[users] row {i}: user_id ไม่ใช่ UUID: {uid_str}")
                AppUser.objects.create(user_id=uid, **defaults)
            else:
                AppUser.objects.create(**defaults)
            created += 1

    return {
        "users_seen": seen,
        "users_created": created,
        "users_updated": updated,
        "users_renamed": renamed,
    }

def import_artists(ws):
    seen = created = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        name = pick(row, ["name","artist","artist_name"])
        if not name: 
            continue
        _, c = Artist.objects.get_or_create(name=name)
        seen += 1; created += 1 if c else 0
    return {"artists_seen": seen, "artists_created": created}

def import_albums(ws):
    seen = created = updated = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        title = pick(row, ["title","album","album_title"])
        if not title: 
            continue
        artist_name = pick(row, ["artist","artist_name"])
        year = as_int(pick(row, ["year","release_year"]))
        cover_url = pick(row, ["cover_url","cover"])  # โมเดลคุณอาจไม่มี ฟังก์ชันจะข้ามให้
        if not artist_name:
            continue
        artist, _ = Artist.objects.get_or_create(name=artist_name)
        alb, c = Album.objects.get_or_create(artist=artist, title=title, defaults={"year": year})
        if not c:
            changed = False
            if year is not None and getattr(alb, "year", None) != year:
                alb.year = year; changed = True
            if cover_url and hasattr(alb, "cover_url") and getattr(alb, "cover_url", None) != cover_url:
                alb.cover_url = cover_url; changed = True
            if changed: 
                alb.save()
                updated += 1
        else:
            created += 1
        seen += 1
    return {"albums_seen": seen, "albums_created": created, "albums_updated": updated}

def import_genres(ws):
    seen = created = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        name = pick(row, ["name","genre"])
        if not name:
            continue
        _, c = Genre.objects.get_or_create(name=name)
        seen += 1; created += 1 if c else 0
    return {"genres_seen": seen, "genres_created": created}

def import_songs(ws):
    seen = created = updated = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        title = pick(row, ["title","song","song_title"])
        artist_name = pick(row, ["artist","artist_name"])
        album_title = pick(row, ["album","album_title"])
        if not title or not artist_name:
            continue
        artist, _ = Artist.objects.get_or_create(name=artist_name)
        album = None
        if album_title:
            album, _ = Album.objects.get_or_create(artist=artist, title=album_title)

        obj, c = Song.objects.get_or_create(title=title, artist=artist, album=album)

        # ฟิลด์เสริม (มีในโมเดลค่อยเซฟ)
        for attr, val in [
            ("duration_sec", as_int(pick(row, ["duration","duration_sec","length_sec"]))),
            ("platform", pick(row, ["platform"])),
            ("external_id", pick(row, ["external_id","ext_id"])),
            ("is_active", as_bool(pick(row, ["is_active","active"]), True)),
            ("lyrics", pick(row, ["lyrics","lyric"])),
        ]:
            if hasattr(obj, attr) and val not in (None, "") and getattr(obj, attr, None) != val:
                setattr(obj, attr, val)

        created_at = parse_dt(pick(row, ["created_at","added_at","created"]))
        if hasattr(obj, "created_at") and created_at and getattr(obj, "created_at", None) in (None, ""):
            obj.created_at = created_at

        # บันทึกถ้ามีการเปลี่ยน
        obj.save()

        seen += 1; created += 1 if c else 0; updated += 0 if c else 1
    return {"songs_seen": seen, "songs_created": created, "songs_updated": updated}

def import_song_genres(ws):
    seen = created = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        title = pick(row, ["song_title","title","song"])
        artist_name = pick(row, ["artist","artist_name"])
        genre_name = pick(row, ["genre","genre_name","name"])
        if not title or not artist_name or not genre_name:
            continue
        artist, _ = Artist.objects.get_or_create(name=artist_name)
        try:
            song = Song.objects.get(title=title, artist=artist)
        except Song.DoesNotExist:
            continue
        genre, _ = Genre.objects.get_or_create(name=genre_name)
        _, c = SongGenre.objects.get_or_create(song=song, genre=genre)
        seen += 1; created += 1 if c else 0
    return {"song_genres_seen": seen, "song_genres_created": created}

def import_emotions(ws):
    if not HAS_EMOTION:
        return {"emotions_seen": 0, "emotions_created": 0}
    seen = created = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        name = pick(row, ["name","emotion"])
        if not name:
            continue
        _, c = Emotion.objects.get_or_create(name=name)
        seen += 1; created += 1 if c else 0
    return {"emotions_seen": seen, "emotions_created": created}

def import_song_emotions(ws):
    if not HAS_EMOTION:
        return {"song_emotions_seen": 0, "song_emotions_created": 0}
    from decimal import Decimal, InvalidOperation
    seen = created = 0
    for i, row in enumerate(sheet_dicts(ws), start=2):
        title = pick(row, ["song_title","title","song"])
        artist_name = pick(row, ["artist","artist_name"])
        emotion_name = pick(row, ["emotion","emotion_name"])
        conf = pick(row, ["confidence","conf"])
        source = pick(row, ["source"], "manual").lower()
        if not title or not artist_name or not emotion_name:
            continue
        artist, _ = Artist.objects.get_or_create(name=artist_name)
        try:
            song = Song.objects.get(title=title, artist=artist)
        except Song.DoesNotExist:
            continue
        emo, _ = Emotion.objects.get_or_create(name=emotion_name)
        try:
            conf_val = Decimal(str(conf)) if conf not in ("", None) else None
        except InvalidOperation:
            conf_val = None
        _, c = SongEmotion.objects.update_or_create(
            song=song, emotion=emo,
            defaults={
                "confidence": conf_val or 0.800,
                "source": source if source in {"ml","rule","manual"} else "manual",
            }
        )
        seen += 1; created += 1 if c else 0
    return {"song_emotions_seen": seen, "song_emotions_created": created}


# -------------------- command --------------------
SHEET_FUNCS = {
    "users": import_users,
    "artists": import_artists,
    "albums": import_albums,
    "songs": import_songs,
    "genres": import_genres,
    "song_genres": import_song_genres,
    "emotions": import_emotions,
    "song_emotions": import_song_emotions,
}

class Command(BaseCommand):
    help = "Import supported sheets from an Excel workbook (.xlsx). Sheets: " + ", ".join(SHEET_FUNCS.keys())

    def add_arguments(self, parser):
        parser.add_argument("xlsx_path", help="Path to .xlsx file")
        parser.add_argument("--only", nargs="*", help="Import only these sheets (e.g. users artists songs)")
        parser.add_argument("--dry-run", action="store_true", help="Parse only, no DB writes")

    @transaction.atomic
    def handle(self, *args, **opts):
        path = opts["xlsx_path"]
        only = set([s.lower() for s in (opts["only"] or [])])
        dry  = opts["dry_run"]

        try:
            wb = load_workbook(filename=path, data_only=True, read_only=True)
        except FileNotFoundError:
            raise CommandError(f"ไม่พบไฟล์: {path}")
        except Exception as e:
            raise CommandError(f"เปิดไฟล์ไม่ได้: {e}")

        # ลำดับที่ปลอดภัย (users -> artists -> albums -> songs -> genres -> song_genres -> emotions -> song_emotions)
        order = [k for k in ["users","artists","albums","songs","genres","song_genres","emotions","song_emotions"] if k in SHEET_FUNCS]
        summary = {}
        for name in order:
            if only and name not in only:
                continue
            if name not in wb.sheetnames:
                continue
            ws = wb[name]
            res = SHEET_FUNCS[name](ws)
            summary.update(res)

        lines = ["=== Import summary ==="] + [f"{k}: {summary[k]}" for k in sorted(summary.keys())]
        out = "\n".join(lines)

        if dry:
            self.stdout.write(self.style.WARNING("[DRY-RUN] ไม่ได้เขียน DB"))
            self.stdout.write(out)
            transaction.set_rollback(True)
            return

        self.stdout.write(self.style.SUCCESS(out))
